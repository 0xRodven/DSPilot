#!/usr/bin/env python3
"""
Build final outputs (Excel + HTML map) merging:
- Annuaire Entreprises data (company identity, dirigeant, address, scoring)
- Enrichment JSON (website, phones, emails, linkedin, procedure)

Usage:
  python3 build-final.py [--enriched /tmp/dsp-idf-enriched.json]

Outputs (idempotent):
  /Users/ousmane/Desktop/leads-DSPilot-IDF.xlsx
  /Users/ousmane/Desktop/leads-DSPilot-IDF.html
  /Users/ousmane/wiki/wiki/leads-DSPilot-IDF.xlsx (mirror)
"""

from __future__ import annotations

import argparse
import json
import sys
import time
from pathlib import Path
from typing import Any

# Reuse functions from source-dsp-idf.py
sys.path.insert(0, str(Path(__file__).parent))
import importlib.util

_spec = importlib.util.spec_from_file_location(
    "src_dsp", str(Path(__file__).parent / "source-dsp-idf.py")
)
src_dsp = importlib.util.module_from_spec(_spec)
sys.modules["src_dsp"] = src_dsp  # required for @dataclass to find module
_spec.loader.exec_module(src_dsp)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--enriched", default="/tmp/dsp-idf-enriched.json")
    args = parser.parse_args()

    # 1. Pull leads via API (idempotent, ~10s)
    print("[1/5] Pull leads via Annuaire Entreprises API…", file=sys.stderr)
    raw_companies: list[dict[str, Any]] = []
    for dept in src_dsp.DEPARTEMENTS_IDF:
        for raw in src_dsp.fetch_dept(dept):
            raw_companies.append(raw)

    # 2. Dedup + parse
    print("[2/5] Dedup + parse + score…", file=sys.stderr)
    seen: set[str] = set()
    companies: list[Any] = []
    for raw in raw_companies:
        siren = raw.get("siren")
        if not siren or siren in seen:
            continue
        seen.add(siren)
        c = src_dsp.parse_company(raw)
        if c is not None:
            companies.append(src_dsp.score_company(c))
    companies.sort(key=lambda c: (-c.score, c.nearest_station_km))

    # 3. Load enriched data
    enriched_map: dict[str, dict[str, Any]] = {}
    enriched_path = Path(args.enriched)
    if enriched_path.exists():
        with open(enriched_path) as f:
            data = json.load(f)
        enriched_map = {e["siren"]: e for e in data}
        print(f"[3/5] Loaded enrichment for {len(enriched_map)} sirens", file=sys.stderr)
    else:
        print(f"[3/5] No enrichment file at {enriched_path} — skipping enrichment merge", file=sys.stderr)

    # 4. Stitch enriched into Company objects (mutate)
    for c in companies:
        e = enriched_map.get(c.siren, {})
        c.website = e.get("website") or ""
        c.domain = e.get("domain") or ""
        c.phones = e.get("phones") or []
        c.emails = e.get("emails") or []
        c.linkedin = e.get("linkedin") or ""
        c.procedure = e.get("procedure") or ""

    # 5. Write outputs
    print("[4/5] Writing Excel…", file=sys.stderr)
    desktop_xlsx = Path("/Users/ousmane/Desktop/leads-DSPilot-IDF.xlsx")
    wiki_xlsx = Path("/Users/ousmane/wiki/wiki/leads-DSPilot-IDF.xlsx")
    write_excel_v2(companies, desktop_xlsx)
    write_excel_v2(companies, wiki_xlsx)

    print("[5/5] Writing HTML interactive map…", file=sys.stderr)
    desktop_html = Path("/Users/ousmane/Desktop/leads-DSPilot-IDF.html")
    write_html_map_v2(companies, desktop_html)

    # Stats
    n = len(companies)
    n_phone = sum(1 for c in companies if getattr(c, "phones", None))
    n_email = sum(1 for c in companies if getattr(c, "emails", None))
    n_web = sum(1 for c in companies if getattr(c, "website", ""))
    n_proc = sum(1 for c in companies if getattr(c, "procedure", ""))
    n_t1 = sum(1 for c in companies if c.score >= 14)
    print(
        f"\n=== STATS ===\n"
        f"Total leads      : {n}\n"
        f"Tier 1 (score≥14): {n_t1}\n"
        f"Avec website     : {n_web} ({n_web*100//n}%)\n"
        f"Avec téléphone   : {n_phone} ({n_phone*100//n}%)\n"
        f"Avec email       : {n_email} ({n_email*100//n}%)\n"
        f"En procédure     : {n_proc} ({n_proc*100//n}%)\n"
        f"\nFichiers générés:\n"
        f"  {desktop_xlsx}\n"
        f"  {desktop_html}\n",
        file=sys.stderr,
    )
    return 0


# ===== Excel v2 (with enrichment cols) =====

def _quote_url(s: str) -> str:
    from urllib.parse import quote_plus
    return quote_plus(s)


HEADERS_V2 = [
    "Score", "Tier", "Stage", "Procédure",
    "SIREN", "Société", "Forme", "Création", "Âge", "Effectif", "Catégorie", "Étabts",
    "Dirigeant Prénom", "Dirigeant Nom", "Fonction", "Année naissance", "Âge dirigeant", "Approche", "Co-dirigeants",
    "Adresse siège", "Code postal", "Commune", "Dept",
    "Lat", "Lng", "Station Amazon", "Distance (km)",
    # Enrichissement
    "Téléphone(s)", "Email(s)", "Site web", "Domain", "LinkedIn",
    # Action / liens
    "Recherche tel (Google)", "Google Maps", "LinkedIn société (rech.)", "LinkedIn dirigeant (rech.)",
    "Annuaire", "Société.com", "MAJ INSEE",
    # Score detail
    "Détail score", "Notes Ousmane",
]


def _company_row_v2(c) -> list[Any]:
    approche = ""
    if c.dirigeant_age is not None:
        approche = "Tutoyer V1" if c.dirigeant_age < 45 else "Vouvoyer V2"

    if c.score >= 14:
        tier_label = "T1"
    elif c.score >= 11:
        tier_label = "T2"
    elif c.score >= 8:
        tier_label = "T3"
    elif c.score >= 5:
        tier_label = "T4"
    else:
        tier_label = "—"

    proc = getattr(c, "procedure", "") or ""

    annuaire = f"https://annuaire-entreprises.data.gouv.fr/entreprise/{c.siren}"
    societe_com = f"https://www.societe.com/cgi-bin/search?champs={c.siren}"
    nom_q = f'"{c.nom}" "{c.commune}" telephone'
    google_tel = f"https://www.google.com/search?q={_quote_url(nom_q)}"
    li_co = f"https://www.google.com/search?q={_quote_url('site:linkedin.com/company ' + chr(34) + c.nom + chr(34))}"
    li_dir = ""
    if c.dirigeant_prenom and c.dirigeant_nom:
        li_dir = f"https://www.google.com/search?q={_quote_url(f'site:linkedin.com/in {chr(34)}{c.dirigeant_prenom} {c.dirigeant_nom}{chr(34)} {chr(34)}{c.commune}{chr(34)}')}"
    if c.latitude and c.longitude:
        gmaps = f"https://www.google.com/maps/search/?api=1&query={c.latitude},{c.longitude}"
    else:
        gmaps = f"https://www.google.com/maps/search/?api=1&query={_quote_url(c.adresse)}"

    phones = " | ".join(getattr(c, "phones", []) or [])
    emails = " | ".join(getattr(c, "emails", []) or [])
    website = getattr(c, "website", "") or ""
    domain = getattr(c, "domain", "") or ""
    linkedin = getattr(c, "linkedin", "") or ""

    # Stage = "Mort" si en procédure
    stage = "Mort - skip" if proc else "Lead"

    return [
        c.score, tier_label, stage, proc,
        c.siren, c.nom, c.forme_juridique_libelle, c.date_creation, c.age_annees, c.tranche_effectif_libelle, c.categorie_entreprise, c.nb_etablissements,
        c.dirigeant_prenom, c.dirigeant_nom, c.dirigeant_fonction, c.dirigeant_annee_naissance, c.dirigeant_age, approche,
        " | ".join(c.other_dirigeants_pp),
        c.adresse, c.code_postal, c.commune, c.departement,
        c.latitude, c.longitude, c.nearest_station_code, c.nearest_station_km if c.nearest_station_km < 999 else None,
        phones, emails, website, domain, linkedin,
        google_tel, gmaps, li_co, li_dir,
        annuaire, societe_com, c.date_maj_insee,
        " ; ".join(c.score_details), "",
    ]


# Index 1-based des cols URLs cliquables dans HEADERS_V2 (actions = label cliquable)
ACTION_LINK_COLS = {
    33: "🔍 Tel Google",
    34: "📍 Maps",
    35: "💼 LkdN société",
    36: "👤 LkdN dirigeant",
    37: "🏢 Annuaire",
    38: "🔎 Société.com",
}
# Cols où l'URL est conservée mais cliquable (Site web, LinkedIn directs)
URL_AS_IS_COLS = {30, 32}  # Site web, LinkedIn (extracted from scrape)


def _style_sheet_v2(ws):
    from openpyxl.styles import Alignment, Font, PatternFill
    HEADER_FILL = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    HEADER_FONT = Font(color="F8FAFC", bold=True, size=11)
    LINK_FONT = Font(color="2563EB", underline="single", size=11)
    PROC_FILL = PatternFill(start_color="450A0A", end_color="450A0A", fill_type="solid")  # rouge foncé
    PROC_FONT = Font(color="FECACA", bold=True, size=11)

    # Header
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "F2"

    # Widths
    widths = {
        "A": 6, "B": 6, "C": 12, "D": 22,
        "E": 11, "F": 32, "G": 8, "H": 11, "I": 5, "J": 9, "K": 6, "L": 6,
        "M": 14, "N": 16, "O": 22, "P": 8, "Q": 6, "R": 13, "S": 28,
        "T": 42, "U": 8, "V": 18, "W": 5,
        "X": 11, "Y": 11, "Z": 8, "AA": 8,
        "AB": 22, "AC": 30, "AD": 30, "AE": 18, "AF": 28,
        "AG": 14, "AH": 12, "AI": 14, "AJ": 14,
        "AK": 12, "AL": 12, "AM": 11,
        "AN": 60, "AO": 25,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Hyperlinks + tier coloring
    for row_idx in range(2, ws.max_row + 1):
        # Procedure col D (index 4) → fill row red if non-empty
        proc = ws.cell(row=row_idx, column=4).value
        is_proc = bool(proc and str(proc).strip())

        # Score color
        sv = ws.cell(row=row_idx, column=1).value or 0
        try:
            sv = int(sv)
        except (TypeError, ValueError):
            sv = 0

        if is_proc:
            row_fill = PROC_FILL
            row_font_color = "FECACA"
        elif sv >= 14:
            row_fill = PatternFill(start_color="A7F3D0", end_color="A7F3D0", fill_type="solid")
            row_font_color = None
        elif sv >= 11:
            row_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
            row_font_color = None
        elif sv >= 8:
            row_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            row_font_color = None
        elif sv >= 5:
            row_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            row_font_color = None
        else:
            row_fill = None
            row_font_color = None

        if row_fill:
            for col_idx in range(1, len(HEADERS_V2) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = row_fill
                if row_font_color and col_idx not in ACTION_LINK_COLS and col_idx not in URL_AS_IS_COLS:
                    cell.font = Font(color=row_font_color, size=11, bold=is_proc)

        # Action URLs (Google search, etc) → label cliquable
        for col_idx, label in ACTION_LINK_COLS.items():
            cell = ws.cell(row=row_idx, column=col_idx)
            url = cell.value
            if url and isinstance(url, str) and url.startswith("http"):
                cell.hyperlink = url
                cell.value = label
                cell.font = LINK_FONT

        # URL as-is (Site web, LinkedIn) → keep URL, make clickable + style
        for col_idx in URL_AS_IS_COLS:
            cell = ws.cell(row=row_idx, column=col_idx)
            url = cell.value
            if url and isinstance(url, str) and url.startswith("http"):
                cell.hyperlink = url
                cell.font = LINK_FONT


def write_excel_v2(companies: list, out_path: Path) -> None:
    from openpyxl import Workbook

    wb = Workbook()

    # Filtrer non-procédure pour les Tier sheets
    alive = [c for c in companies if not getattr(c, "procedure", "")]
    dead = [c for c in companies if getattr(c, "procedure", "")]

    tier1 = [c for c in alive if c.score >= 14]
    tier2 = [c for c in alive if 11 <= c.score < 14]
    tier3 = [c for c in alive if 8 <= c.score < 11]
    tier4 = [c for c in alive if 5 <= c.score < 8]

    sheets = [
        ("Tier 1 - Action immédiate", tier1),
        ("Tier 2 - High priority", tier2),
        ("Tier 3 - Secondary", tier3),
        ("Tier 4 - Exploration", tier4),
        ("All - Vue complète", companies),
        ("⚠️ En procédure (skip)", dead),
    ]

    for i, (title, leads) in enumerate(sheets):
        if i == 0:
            ws = wb.active
            if ws is None:
                ws = wb.create_sheet()
            ws.title = title
        else:
            ws = wb.create_sheet(title)
        ws.append(HEADERS_V2)
        for c in leads:
            ws.append(_company_row_v2(c))
        if leads:
            _style_sheet_v2(ws)

    # Stations sheet
    ws_st = wb.create_sheet("Stations Amazon")
    ws_st.append(["Code", "Localisation", "Lat", "Lng"])
    for code, (label, lat, lng) in src_dsp.AMAZON_STATIONS.items():
        ws_st.append([code, label, lat, lng])
    from openpyxl.styles import Font, PatternFill
    for cell in ws_st[1]:
        cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        cell.font = Font(color="F8FAFC", bold=True)
    ws_st.column_dimensions["A"].width = 10
    ws_st.column_dimensions["B"].width = 45
    ws_st.column_dimensions["C"].width = 12
    ws_st.column_dimensions["D"].width = 12

    # Méthodologie
    ws_m = wb.create_sheet("Méthodologie")
    info = [
        ["DSPilot — Leads DSP Amazon Île-de-France"],
        [""],
        ["Source primaire", "Annuaire Entreprises (recherche-entreprises.api.gouv.fr) — État FR, gratuit"],
        ["Source enrichissement", "Domain guess + scrape site web société (homepage + /contact + /mentions-legales)"],
        ["Source procédure", "société.com (détection liquidation/redressement par scrape)"],
        ["Date extraction", time.strftime("%Y-%m-%d %H:%M")],
        [""],
        ["Filtres appliqués"],
        ["", "APE 49.41B", "Transports routiers de fret de proximité"],
        ["", "Départements", "75, 77, 78, 91, 92, 93, 94, 95"],
        ["", "Effectif", "10-199 (tranches INSEE 11, 12, 21, 22)"],
        ["", "Catégorie", "PME/ETI uniquement (GE = grandes entreprises exclues)"],
        ["", "État", "Actif"],
        [""],
        ["Scoring (max 20 pts) — voir détail dans col 'Détail score'"],
        ["", "Proximité station Amazon (max 6)"],
        ["", "Programme DSP FR (création post-2018) (max 3)"],
        ["", "Effectif sweet spot 50-99 (max 4)"],
        ["", "Forme juridique SAS/SARL (max 2)"],
        ["", "Dirigeant identifié (max 2)"],
        ["", "Mono-établissement (max 2)"],
        [""],
        ["Lecture des onglets"],
        ["", "Tier 1 (score ≥ 14)", "🟢 Action immédiate — voisins direct station, dirigeant identifié, taille DSP"],
        ["", "Tier 2 (score 11-13)", "🟢 High priority"],
        ["", "Tier 3 (score 8-10)", "🟡 Secondary — vérifier réseau"],
        ["", "Tier 4 (score 5-7)", "🟠 Exploration — probablement non-DSP"],
        ["", "⚠️ En procédure", "Liquidation / Redressement détecté via société.com — NE PAS contacter"],
        [""],
        ["Workflow"],
        ["", "1.", "Ouvre Tier 1 — parcours les 79 lignes top"],
        ["", "2.", "Pour chaque ligne sans tel : clic 'Recherche tel (Google)' → tel trouvé en 5 sec"],
        ["", "3.", "Clic 'Google Maps' → check le quartier"],
        ["", "4.", "Clic 'LinkedIn société/dirigeant' → vérifier le profil"],
        ["", "5.", "Update col 'Stage' : Lead → Enriched → Contacté → Booked → Démo → Won/Lost"],
        ["", "6.", "Pour les ligne en rouge (procédure) : skip, ne perds pas ton temps"],
        [""],
        ["RGPD", "Toutes données publiques (INSEE SIRENE + RNE). Aucune adresse personnelle dirigeant. Conforme outreach B2B."],
    ]
    for row in info:
        ws_m.append(row)
    ws_m.column_dimensions["A"].width = 6
    ws_m.column_dimensions["B"].width = 28
    ws_m.column_dimensions["C"].width = 100
    ws_m["A1"].font = Font(bold=True, size=14)

    wb.save(out_path)


# ===== HTML map v2 =====

def write_html_map_v2(companies: list, out_path: Path) -> None:
    leads_data = []
    for c in companies:
        if c.latitude is None or c.longitude is None:
            continue

        nom_q = f'"{c.nom}" "{c.commune}" telephone'
        gtel = f"https://www.google.com/search?q={_quote_url(nom_q)}"
        li_co = f"https://www.google.com/search?q={_quote_url('site:linkedin.com/company ' + chr(34) + c.nom + chr(34))}"
        li_dir = ""
        if c.dirigeant_prenom and c.dirigeant_nom:
            li_dir = f"https://www.google.com/search?q={_quote_url(f'site:linkedin.com/in {chr(34)}{c.dirigeant_prenom} {c.dirigeant_nom}{chr(34)} {chr(34)}{c.commune}{chr(34)}')}"
        gmaps = f"https://www.google.com/maps/search/?api=1&query={c.latitude},{c.longitude}"
        annuaire = f"https://annuaire-entreprises.data.gouv.fr/entreprise/{c.siren}"
        societecom = f"https://www.societe.com/cgi-bin/search?champs={c.siren}"

        approche = ""
        if c.dirigeant_age is not None:
            approche = "Tutoyer V1" if c.dirigeant_age < 45 else "Vouvoyer V2"

        proc = getattr(c, "procedure", "") or ""
        if proc:
            tier = 0  # mort
        elif c.score >= 14:
            tier = 1
        elif c.score >= 11:
            tier = 2
        elif c.score >= 8:
            tier = 3
        elif c.score >= 5:
            tier = 4
        else:
            tier = 5

        leads_data.append({
            "siren": c.siren,
            "nom": c.nom,
            "lat": c.latitude,
            "lng": c.longitude,
            "score": c.score,
            "tier": tier,
            "procedure": proc,
            "forme": c.forme_juridique_libelle,
            "creation": c.date_creation,
            "age": c.age_annees,
            "effectif": c.tranche_effectif_libelle,
            "etabts": c.nb_etablissements,
            "dirigeant": f"{c.dirigeant_prenom} {c.dirigeant_nom}".strip(),
            "fonction": c.dirigeant_fonction,
            "age_dir": c.dirigeant_age,
            "approche": approche,
            "adresse": c.adresse,
            "cp": c.code_postal,
            "commune": c.commune,
            "dept": c.departement,
            "station": c.nearest_station_code,
            "station_label": c.nearest_station_label,
            "km": round(c.nearest_station_km, 2) if c.nearest_station_km < 999 else None,
            "score_details": " ; ".join(c.score_details),
            "phones": getattr(c, "phones", []) or [],
            "emails": getattr(c, "emails", []) or [],
            "website": getattr(c, "website", "") or "",
            "linkedin": getattr(c, "linkedin", "") or "",
            "gtel": gtel,
            "li_co": li_co,
            "li_dir": li_dir,
            "gmaps": gmaps,
            "annuaire": annuaire,
            "societecom": societecom,
        })

    stations_data = [
        {"code": code, "label": label, "lat": lat, "lng": lng}
        for code, (label, lat, lng) in src_dsp.AMAZON_STATIONS.items()
    ]

    leads_json = json.dumps(leads_data, ensure_ascii=False)
    stations_json = json.dumps(stations_data, ensure_ascii=False)

    html = HTML_TEMPLATE_V2 \
        .replace("__GENERATED__", time.strftime("%Y-%m-%d %H:%M")) \
        .replace("__TOTAL__", str(len(leads_data))) \
        .replace("__LEADS_JSON__", leads_json) \
        .replace("__STATIONS_JSON__", stations_json)

    out_path.write_text(html, encoding="utf-8")


HTML_TEMPLATE_V2 = """<!doctype html>
<html lang="fr">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>DSPilot — Carte leads DSP IDF</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css">
<link rel="stylesheet" href="https://unpkg.com/leaflet.markercluster@1.5.3/dist/MarkerCluster.css">
<link rel="stylesheet" href="https://unpkg.com/leaflet.markercluster@1.5.3/dist/MarkerCluster.Default.css">
<style>
* { box-sizing: border-box; }
body { margin:0; font:14px/1.4 -apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif; color:#111827; }
#app { display:grid; grid-template-columns:340px 1fr; height:100vh; }
#sidebar { background:#0f172a; color:#e2e8f0; padding:16px; overflow-y:auto; }
#sidebar h1 { margin:0 0 4px; font-size:18px; color:#f8fafc; }
#sidebar .meta { color:#94a3b8; font-size:12px; margin-bottom:16px; }
#sidebar h2 { margin:18px 0 8px; font-size:12px; text-transform:uppercase; letter-spacing:.05em; color:#94a3b8; font-weight:700; }
#sidebar .stat { display:flex; justify-content:space-between; padding:6px 0; border-bottom:1px solid #1e293b; font-size:13px; }
#sidebar .stat .v { font-weight:600; color:#f8fafc; }
#sidebar input[type=text], #sidebar select { width:100%; padding:7px 10px; border-radius:6px; border:1px solid #334155; background:#1e293b; color:#f1f5f9; font-size:13px; margin-bottom:8px; }
#sidebar label { display:flex; align-items:center; gap:6px; padding:4px 0; font-size:13px; cursor:pointer; }
#sidebar label input { margin:0; }
#sidebar .tier-pill { display:inline-block; width:10px; height:10px; border-radius:50%; margin-right:6px; vertical-align:middle; }
#sidebar .reset { display:block; width:100%; padding:8px; margin-top:12px; background:#1e293b; color:#e2e8f0; border:1px solid #334155; border-radius:6px; cursor:pointer; font-size:13px; }
#sidebar .reset:hover { background:#334155; }
#sidebar .legend { font-size:11px; color:#64748b; margin-top:14px; padding-top:10px; border-top:1px solid #1e293b; }
#map { height:100vh; }
.popup { font-size:13px; min-width:300px; max-width:380px; }
.popup h3 { margin:0 0 4px; font-size:15px; color:#0f172a; line-height:1.25; }
.popup .siren { color:#64748b; font-size:11px; font-family:ui-monospace,monospace; }
.popup .score { display:inline-block; padding:2px 8px; border-radius:4px; font-weight:700; margin-left:6px; font-size:11px; }
.popup .score.t1 { background:#10b981; color:white; }
.popup .score.t2 { background:#34d399; color:white; }
.popup .score.t3 { background:#fbbf24; color:#78350f; }
.popup .score.t4 { background:#f87171; color:white; }
.popup .score.t5 { background:#cbd5e1; color:#1e293b; }
.popup .score.t0 { background:#7f1d1d; color:white; }
.popup .proc-warn { background:#fef2f2; color:#991b1b; padding:6px 10px; border-radius:5px; margin:8px 0; font-weight:600; border:1px solid #fca5a5; }
.popup .grid { display:grid; grid-template-columns:auto 1fr; gap:4px 12px; margin:8px 0; }
.popup .grid b { color:#475569; font-weight:500; }
.popup .ph-found { background:#dbeafe; border:1px solid #93c5fd; padding:6px 10px; border-radius:5px; margin:8px 0; font-size:12px; }
.popup .ph-found b { color:#1e40af; }
.popup .ph-found a { color:#1e40af; text-decoration:none; font-weight:600; }
.popup .links { display:flex; flex-wrap:wrap; gap:5px; margin-top:10px; padding-top:10px; border-top:1px solid #e2e8f0; }
.popup .links a { display:inline-block; padding:5px 10px; background:#2563eb; color:white; text-decoration:none; border-radius:5px; font-size:11px; font-weight:600; }
.popup .links a:hover { background:#1d4ed8; }
.popup .links a.secondary { background:#64748b; }
.popup .links a.secondary:hover { background:#475569; }
.popup .links a.success { background:#059669; }
.leaflet-popup-content { margin:14px; }
</style>
</head>
<body>
<div id="app">
<div id="sidebar">
  <h1>DSPilot — Carte leads</h1>
  <div class="meta">__GENERATED__ • __TOTAL__ leads • Source : Annuaire Entreprises FR</div>

  <h2>🔍 Recherche libre</h2>
  <input type="text" id="search" placeholder="Société, dirigeant, ville, SIREN…">

  <h2>Tiers (score DSP-fit)</h2>
  <label><input type="checkbox" class="tier" value="1" checked><span class="tier-pill" style="background:#10b981"></span>T1 — Action immédiate (≥14)</label>
  <label><input type="checkbox" class="tier" value="2" checked><span class="tier-pill" style="background:#34d399"></span>T2 — High priority (11-13)</label>
  <label><input type="checkbox" class="tier" value="3"><span class="tier-pill" style="background:#fbbf24"></span>T3 — Secondary (8-10)</label>
  <label><input type="checkbox" class="tier" value="4"><span class="tier-pill" style="background:#f87171"></span>T4 — Exploration (5-7)</label>
  <label><input type="checkbox" class="tier" value="5"><span class="tier-pill" style="background:#cbd5e1"></span>Bruit (&lt;5)</label>

  <h2>Filtres</h2>
  <label><input type="checkbox" id="hide-proc" checked>Cacher en procédure (mort)</label>
  <label><input type="checkbox" id="only-phone">Uniquement avec téléphone</label>
  <label><input type="checkbox" id="only-website">Uniquement avec site web</label>

  <h2>Station Amazon</h2>
  <select id="station-filter"><option value="">Toutes</option></select>

  <h2>Effectif</h2>
  <select id="effectif-filter">
    <option value="">Tous</option>
    <option value="10-19">10-19</option>
    <option value="20-49">20-49</option>
    <option value="50-99">50-99</option>
    <option value="100-199">100-199</option>
  </select>

  <h2>Département</h2>
  <select id="dept-filter">
    <option value="">Tous</option>
    <option value="75">75 Paris</option>
    <option value="77">77 Seine-et-Marne</option>
    <option value="78">78 Yvelines</option>
    <option value="91">91 Essonne</option>
    <option value="92">92 Hauts-de-Seine</option>
    <option value="93">93 Seine-Saint-Denis</option>
    <option value="94">94 Val-de-Marne</option>
    <option value="95">95 Val-d'Oise</option>
  </select>

  <h2>Visibles</h2>
  <div class="stat"><span>Total affichés</span><span class="v" id="count-visible">0</span></div>
  <div class="stat"><span>Avec téléphone</span><span class="v" id="count-phone">0</span></div>
  <div class="stat"><span>Avec site web</span><span class="v" id="count-web">0</span></div>

  <button class="reset" onclick="resetFilters()">Reset filtres</button>

  <div class="legend">
    🟣 Pin violet = station Amazon Logistics<br>
    Cliquer un lead pour voir détail + liens d'action
  </div>
</div>
<div id="map"></div>
</div>

<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<script src="https://unpkg.com/leaflet.markercluster@1.5.3/dist/leaflet.markercluster.js"></script>
<script>
const LEADS = __LEADS_JSON__;
const STATIONS = __STATIONS_JSON__;

const tierColor = {0:"#7f1d1d",1:"#10b981",2:"#34d399",3:"#fbbf24",4:"#f87171",5:"#cbd5e1"};
const tierBorder = {0:"#450a0a",1:"#065f46",2:"#047857",3:"#b45309",4:"#991b1b",5:"#475569"};

const map = L.map("map", { preferCanvas:true }).setView([48.8566, 2.3522], 10);
L.tileLayer("https://{s}.basemaps.cartocdn.com/light_all/{z}/{x}/{y}{r}.png", {
  attribution:"&copy; OpenStreetMap & CARTO", maxZoom:19
}).addTo(map);

STATIONS.forEach(s => {
  L.marker([s.lat, s.lng], {
    icon: L.divIcon({
      className:"",
      html:`<div style="background:#a855f7;border:3px solid white;border-radius:50%;width:22px;height:22px;box-shadow:0 0 0 2px #a855f7,0 1px 4px rgba(0,0,0,.3)"></div>`,
      iconSize:[22,22], iconAnchor:[11,11]
    }),
    zIndexOffset: 1000
  }).addTo(map).bindPopup(`<b>Station Amazon ${s.code}</b><br>${s.label}`);
});

const stationFilter = document.getElementById("station-filter");
[...new Set(LEADS.map(l => l.station).filter(Boolean))].sort().forEach(code => {
  const opt = document.createElement("option");
  opt.value = code; opt.textContent = code;
  stationFilter.appendChild(opt);
});

const cluster = L.markerClusterGroup({ maxClusterRadius:50, spiderfyOnMaxZoom:true, showCoverageOnHover:false });
const markers = [];

LEADS.forEach(l => {
  const color = tierColor[l.tier] || "#cbd5e1";
  const border = tierBorder[l.tier] || "#475569";
  const size = l.tier === 1 ? 14 : l.tier === 2 ? 12 : 10;
  const m = L.circleMarker([l.lat, l.lng], {
    radius: size/2, fillColor: color, color: border, weight: 2, fillOpacity: 0.85,
  });
  m.leadData = l;
  m.bindPopup(() => buildPopup(l), { maxWidth: 400 });
  markers.push(m);
});

function buildPopup(l) {
  const procWarn = l.procedure ? `<div class="proc-warn">⚠️ ${escapeHtml(l.procedure)} — ne pas contacter</div>` : "";

  const phoneRow = l.phones && l.phones.length
    ? `<div class="ph-found"><b>📞 Téléphone trouvé :</b><br>${l.phones.map(p => `<a href="tel:${p}">${escapeHtml(p)}</a>`).join(" • ")}</div>`
    : "";

  const emailRow = l.emails && l.emails.length
    ? `<div class="ph-found"><b>✉️ Email :</b><br>${l.emails.map(e => `<a href="mailto:${e}">${escapeHtml(e)}</a>`).join(" • ")}</div>`
    : "";

  let websiteRow = "";
  if (l.website) {
    let host = l.website;
    try { host = new URL(l.website).hostname; } catch(e) {}
    websiteRow = `<div class="grid"><b>Site web</b><span><a href="${l.website}" target="_blank">${escapeHtml(host)}</a></span></div>`;
  }

  const dirig = l.dirigeant
    ? `<div class="grid"><b>Dirigeant</b><span>${escapeHtml(l.dirigeant)}</span><b>Fonction</b><span>${escapeHtml(l.fonction || "")}</span><b>Approche</b><span><b>${l.approche || "?"}</b> (${l.age_dir ?? "?"} ans)</span></div>`
    : `<div class="grid"><b>Dirigeant</b><span style="color:#94a3b8">Non identifié dans RNE</span></div>`;

  const km = l.km != null ? `${l.km} km` : "?";
  const links = [
    !l.phones || !l.phones.length ? `<a href="${l.gtel}" target="_blank">📞 Trouver tel</a>` : "",
    `<a href="${l.gmaps}" target="_blank">📍 Google Maps</a>`,
    `<a href="${l.li_co}" target="_blank" class="secondary">LinkedIn société</a>`,
    l.li_dir ? `<a href="${l.li_dir}" target="_blank" class="secondary">LinkedIn dirigeant</a>` : "",
    `<a href="${l.annuaire}" target="_blank" class="secondary">Annuaire</a>`,
    `<a href="${l.societecom}" target="_blank" class="secondary">Société.com</a>`,
  ].filter(Boolean).join("");

  return `<div class="popup">
    <h3>${escapeHtml(l.nom)} <span class="score t${l.tier}">${l.score}</span></h3>
    <div class="siren">SIREN ${l.siren} • ${l.forme || "?"} • Créée ${l.creation || "?"} (${l.age || "?"} ans) • ${l.etabts || 1} étabt(s)</div>
    ${procWarn}
    ${phoneRow}
    ${emailRow}
    <div class="grid" style="margin-top:8px">
      <b>Effectif</b><span>${l.effectif || "?"} salariés</span>
      <b>Adresse</b><span>${escapeHtml(l.adresse)}</span>
      <b>Station</b><span>${l.station || "?"} (${km})</span>
    </div>
    ${websiteRow}
    ${dirig}
    <div class="grid"><b>Score</b><span style="font-size:11px;color:#64748b">${escapeHtml(l.score_details || "")}</span></div>
    <div class="links">${links}</div>
  </div>`;
}

function escapeHtml(s) { return String(s ?? "").replace(/[&<>"']/g, c => ({"&":"&amp;","<":"&lt;",">":"&gt;","\\"":"&quot;","'":"&#39;"})[c]); }

const filters = {
  tiers: new Set([1,2]),
  search: "",
  station: "", effectif: "", dept: "",
  hideProc: true, onlyPhone: false, onlyWebsite: false,
};

function applyFilters() {
  const q = filters.search.toLowerCase();
  cluster.clearLayers();
  let visible = 0, withPhone = 0, withWeb = 0;
  markers.forEach(m => {
    const l = m.leadData;
    if (filters.hideProc && l.procedure) return;
    if (!filters.hideProc && l.tier === 0 && !filters.tiers.has(0)) {
      // si on affiche pas hideProc et tier 0 pas dans tiers, skip
    }
    if (l.tier !== 0 && !filters.tiers.has(l.tier)) return;
    if (l.tier === 0 && filters.hideProc) return;  // procédure cachée
    if (filters.station && l.station !== filters.station) return;
    if (filters.effectif && l.effectif !== filters.effectif) return;
    if (filters.dept && l.dept !== filters.dept) return;
    if (filters.onlyPhone && (!l.phones || !l.phones.length)) return;
    if (filters.onlyWebsite && !l.website) return;
    if (q) {
      const hay = `${l.nom} ${l.dirigeant} ${l.commune} ${l.siren} ${l.adresse}`.toLowerCase();
      if (!hay.includes(q)) return;
    }
    cluster.addLayer(m);
    visible++;
    if (l.phones && l.phones.length) withPhone++;
    if (l.website) withWeb++;
  });
  if (!map.hasLayer(cluster)) map.addLayer(cluster);
  document.getElementById("count-visible").textContent = visible;
  document.getElementById("count-phone").textContent = withPhone;
  document.getElementById("count-web").textContent = withWeb;
}

document.querySelectorAll(".tier").forEach(cb => cb.addEventListener("change", () => {
  const v = parseInt(cb.value, 10);
  if (cb.checked) filters.tiers.add(v); else filters.tiers.delete(v);
  applyFilters();
}));
document.getElementById("search").addEventListener("input", e => { filters.search = e.target.value; applyFilters(); });
document.getElementById("station-filter").addEventListener("change", e => { filters.station = e.target.value; applyFilters(); });
document.getElementById("effectif-filter").addEventListener("change", e => { filters.effectif = e.target.value; applyFilters(); });
document.getElementById("dept-filter").addEventListener("change", e => { filters.dept = e.target.value; applyFilters(); });
document.getElementById("hide-proc").addEventListener("change", e => { filters.hideProc = e.target.checked; applyFilters(); });
document.getElementById("only-phone").addEventListener("change", e => { filters.onlyPhone = e.target.checked; applyFilters(); });
document.getElementById("only-website").addEventListener("change", e => { filters.onlyWebsite = e.target.checked; applyFilters(); });

function resetFilters() {
  filters.tiers = new Set([1,2]);
  filters.search = ""; filters.station = ""; filters.effectif = ""; filters.dept = "";
  filters.hideProc = true; filters.onlyPhone = false; filters.onlyWebsite = false;
  document.querySelectorAll(".tier").forEach(cb => { cb.checked = [1,2].includes(parseInt(cb.value,10)); });
  document.getElementById("search").value = "";
  document.getElementById("station-filter").value = "";
  document.getElementById("effectif-filter").value = "";
  document.getElementById("dept-filter").value = "";
  document.getElementById("hide-proc").checked = true;
  document.getElementById("only-phone").checked = false;
  document.getElementById("only-website").checked = false;
  applyFilters();
}

applyFilters();
</script>
</body>
</html>"""


if __name__ == "__main__":
    sys.exit(main())
