#!/usr/bin/env python3
"""
Source DSP candidats Amazon France — Île-de-France pilote.

Pipeline :
1. Annuaire Entreprises API (recherche-entreprises.api.gouv.fr) — gratuit, illimité
2. Filtre APE 49.41B + IDF (8 départements) + effectif 10-199 + actif + PME (pas GE)
3. Distance haversine vers stations Amazon Logistics
4. Scoring multi-critères → DSP candidate score
5. Output Excel multi-sheets dans wiki/leads-DSPilot.xlsx

Source : 100% data publique RGPD-compliant (INSEE SIRENE + RNE)
"""

from __future__ import annotations

import json
import sys
import time
from dataclasses import dataclass, field
from math import asin, cos, radians, sin, sqrt
from pathlib import Path
from typing import Any, Iterator

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

API_BASE = "https://recherche-entreprises.api.gouv.fr/search"
USER_AGENT = "DSPilot/1.0 (contact: ousmane@dspilot.fr) - pilote DSP IDF"
DEPARTEMENTS_IDF = ["75", "77", "78", "91", "92", "93", "94", "95"]
TRANCHES_EFFECTIF = "11,12,21,22"  # 10-19, 20-49, 50-99, 100-199 salariés

# Stations Amazon Logistics France connues couvrant IDF + zone élargie
AMAZON_STATIONS: dict[str, tuple[str, float, float]] = {
    "DIF1": ("Ivry-sur-Seine (94) — station d'Ousmane", 48.8127, 2.3870),
    "CDG10": ("Clichy (92)", 48.9047, 2.3036),
    "ORY1": ("Saran (45) — hors IDF, sud", 47.9620, 1.8836),
    "LON1": ("Senlis (60) — nord IDF", 49.2000, 2.5800),
    "BIA1": ("Boves (80) — picardie", 49.8400, 2.4000),
    "MLE1": ("Brétigny-sur-Orge (91)", 48.6105, 2.3050),
    "VLY1": ("Vélizy-Villacoublay (78)", 48.7826, 2.1879),
    "FCY1": ("Lieusaint (77)", 48.6263, 2.5532),
}

# Tranches d'effectif INSEE → libellé lisible
TRANCHE_LABELS = {
    "00": "0 salarié",
    "01": "1-2",
    "02": "3-5",
    "03": "6-9",
    "11": "10-19",
    "12": "20-49",
    "21": "50-99",
    "22": "100-199",
    "31": "200-249",
    "32": "250-499",
}

# Mapping forme juridique INSEE (top valeurs courantes)
FORMES_JURIDIQUES = {
    "5710": "SAS",
    "5499": "SARL",
    "5485": "SARL associé unique",
    "5720": "SASU",
    "5202": "SNC",
    "5306": "SCS",
    "1000": "EI",
    "5800": "SE",
    "9220": "Association",
}


@dataclass
class Company:
    siren: str
    nom: str
    departement: str
    code_postal: str
    commune: str
    adresse: str
    latitude: float | None
    longitude: float | None
    date_creation: str
    age_annees: int
    forme_juridique: str
    forme_juridique_libelle: str
    categorie_entreprise: str
    tranche_effectif: str
    tranche_effectif_libelle: str
    nb_etablissements: int
    dirigeant_prenom: str = ""
    dirigeant_nom: str = ""
    dirigeant_fonction: str = ""
    dirigeant_annee_naissance: str = ""
    dirigeant_age: int | None = None
    other_dirigeants_pp: list[str] = field(default_factory=list)
    nearest_station_code: str = ""
    nearest_station_label: str = ""
    nearest_station_km: float = 999.0
    score: int = 0
    score_details: list[str] = field(default_factory=list)
    date_maj_insee: str = ""


def haversine_km(lat1: float, lng1: float, lat2: float, lng2: float) -> float:
    R = 6371.0
    lat1, lng1, lat2, lng2 = map(radians, (lat1, lng1, lat2, lng2))
    dlat, dlng = lat2 - lat1, lng2 - lng1
    a = sin(dlat / 2) ** 2 + cos(lat1) * cos(lat2) * sin(dlng / 2) ** 2
    return 2 * R * asin(sqrt(a))


def fetch_dept(dept: str) -> Iterator[dict[str, Any]]:
    """Stream toutes les sociétés actives APE 49.41B effectif 10-199 du département."""
    page = 1
    per_page = 25
    while True:
        params = {
            "activite_principale": "49.41B",
            "departement": dept,
            "etat_administratif": "A",
            "tranche_effectif_salarie": TRANCHES_EFFECTIF,
            "page": page,
            "per_page": per_page,
        }
        r = requests.get(
            API_BASE,
            params=params,
            headers={"User-Agent": USER_AGENT},
            timeout=15,
        )
        r.raise_for_status()
        data = r.json()
        results = data.get("results", [])
        if not results:
            break
        for item in results:
            yield item
        total_pages = data.get("total_pages", 0)
        if page >= total_pages:
            break
        page += 1
        time.sleep(0.1)  # poli


def parse_company(raw: dict[str, Any]) -> Company | None:
    """Convert raw API result → Company. None si données invalides."""
    siren = raw.get("siren")
    if not siren:
        return None

    siege = raw.get("siege", {}) or {}
    if not siege:
        return None

    # Filtre négatif : grandes entreprises (GE) — pas la cible
    if raw.get("categorie_entreprise") == "GE":
        return None

    date_creation = raw.get("date_creation") or siege.get("date_creation") or ""
    age = 0
    if date_creation:
        try:
            year = int(date_creation[:4])
            age = 2026 - year
        except (ValueError, IndexError):
            pass

    nature = raw.get("nature_juridique") or ""
    forme_libelle = FORMES_JURIDIQUES.get(nature, nature)

    tranche = raw.get("tranche_effectif_salarie") or siege.get("tranche_effectif_salarie") or ""
    tranche_lib = TRANCHE_LABELS.get(tranche, tranche)

    lat_str = siege.get("latitude")
    lng_str = siege.get("longitude")
    lat = float(lat_str) if lat_str else None
    lng = float(lng_str) if lng_str else None

    maj_insee = raw.get("date_mise_a_jour_insee") or siege.get("date_mise_a_jour_insee") or ""
    if maj_insee and "T" in maj_insee:
        maj_insee = maj_insee.split("T")[0]

    company = Company(
        siren=siren,
        nom=raw.get("nom_complet") or raw.get("nom_raison_sociale") or "",
        departement=siege.get("departement") or "",
        code_postal=siege.get("code_postal") or "",
        commune=siege.get("libelle_commune") or "",
        adresse=siege.get("adresse") or "",
        latitude=lat,
        longitude=lng,
        date_creation=date_creation,
        age_annees=age,
        forme_juridique=nature,
        forme_juridique_libelle=forme_libelle,
        categorie_entreprise=raw.get("categorie_entreprise") or "",
        tranche_effectif=tranche,
        tranche_effectif_libelle=tranche_lib,
        nb_etablissements=raw.get("nombre_etablissements") or 0,
        date_maj_insee=maj_insee,
    )

    # Dirigeants — keep PP (personne physique) only, exclure commissaires aux comptes
    dirigeants = raw.get("dirigeants", []) or []
    targets_fonctions = {
        "Président",
        "Président de SAS",
        "Présidente de SAS",
        "Directeur Général",
        "Directrice Générale",
        "Gérant",
        "Gérante",
        "Co-gérant",
        "Co-gérante",
        "Président-Directeur Général",
        "Président du conseil d'administration",
    }
    pp_dirigeants = []
    for d in dirigeants:
        if d.get("type_dirigeant") != "personne physique":
            continue
        qual = d.get("qualite") or ""
        # Exclure commissaires aux comptes
        if "ommissaire" in qual:
            continue
        pp_dirigeants.append(d)

    if pp_dirigeants:
        # Prendre le premier "vrai" dirigeant cible si possible
        primary = next(
            (d for d in pp_dirigeants if any(t in (d.get("qualite") or "") for t in targets_fonctions)),
            pp_dirigeants[0],
        )
        company.dirigeant_prenom = (primary.get("prenoms") or "").title().split(" ")[0]
        company.dirigeant_nom = (primary.get("nom") or "").upper()
        company.dirigeant_fonction = primary.get("qualite") or ""
        annee = primary.get("annee_de_naissance") or ""
        company.dirigeant_annee_naissance = annee
        if annee:
            try:
                company.dirigeant_age = 2026 - int(annee)
            except ValueError:
                pass

        # Autres dirigeants PP
        others = [d for d in pp_dirigeants if d is not primary]
        company.other_dirigeants_pp = [
            f"{(d.get('prenoms') or '').title().split(' ')[0]} {(d.get('nom') or '').upper()} ({d.get('qualite') or ''})"
            for d in others
        ]

    # Distance vers la station Amazon la plus proche
    if lat and lng:
        best_code = ""
        best_lib = ""
        best_km = 999.0
        for code, (label, slat, slng) in AMAZON_STATIONS.items():
            d = haversine_km(lat, lng, slat, slng)
            if d < best_km:
                best_km, best_code, best_lib = d, code, label
        company.nearest_station_code = best_code
        company.nearest_station_label = best_lib
        company.nearest_station_km = round(best_km, 2)

    return company


def score_company(c: Company) -> Company:
    """Score 0-20 — heuristique DSP Amazon candidate (resserrée)."""
    score = 0
    details = []

    # Proximité station Amazon (poids fort, resserré)
    if c.nearest_station_km <= 2:
        score += 6
        details.append(f"+6 station <= 2 km ({c.nearest_station_code})")
    elif c.nearest_station_km <= 5:
        score += 4
        details.append(f"+4 station <= 5 km ({c.nearest_station_code})")
    elif c.nearest_station_km <= 10:
        score += 2
        details.append(f"+2 station <= 10 km ({c.nearest_station_code})")
    elif c.nearest_station_km <= 20:
        score += 1
        details.append(f"+1 station <= 20 km ({c.nearest_station_code})")

    # Date création alignée programme DSP Amazon FR (lancé 2018)
    year_creation = 0
    if c.date_creation:
        try:
            year_creation = int(c.date_creation[:4])
        except (ValueError, IndexError):
            pass
    if year_creation >= 2018:
        score += 3
        details.append(f"+3 création {year_creation} (post-DSP launch FR)")
    elif year_creation >= 2015:
        score += 1
        details.append(f"+1 création {year_creation}")

    # Effectif typique DSP (30-100 livreurs)
    if c.tranche_effectif == "21":  # 50-99
        score += 4
        details.append("+4 effectif 50-99 (sweet spot DSP)")
    elif c.tranche_effectif == "12":  # 20-49
        score += 3
        details.append("+3 effectif 20-49 (taille DSP)")
    elif c.tranche_effectif == "11":  # 10-19
        score += 1
        details.append("+1 effectif 10-19 (sous-DSP ou jeune DSP)")
    elif c.tranche_effectif == "22":  # 100-199
        score += 1
        details.append("+1 effectif 100-199 (peut-etre multi-station)")

    # Forme juridique programme DSP exige société commerciale
    if c.forme_juridique_libelle in {"SAS", "SARL", "SARL associé unique", "SASU"}:
        score += 2
        details.append(f"+2 forme {c.forme_juridique_libelle}")

    # Dirigeant identifié
    if c.dirigeant_prenom and c.dirigeant_nom:
        score += 2
        details.append(f"+2 dirigeant identifie")

    # 1 seul établissement = vrai DSP indépendant
    if c.nb_etablissements == 1:
        score += 2
        details.append("+2 mono-etablissement (DSP independant)")
    elif c.nb_etablissements <= 3:
        score += 1
        details.append("+1 2-3 etablissements")

    c.score = score
    c.score_details = details
    return c


HEADERS = [
    "Score",
    "Stage",
    "SIREN",
    "Société",
    "Forme",
    "Création",
    "Âge",
    "Effectif",
    "Catégorie",
    "Étabts",
    "Dirigeant Prénom",
    "Dirigeant Nom",
    "Fonction",
    "Année naissance",
    "Âge dirigeant",
    "Approche",
    "Co-dirigeants PP",
    "Adresse siège",
    "Code postal",
    "Commune",
    "Dept",
    "Lat",
    "Lng",
    "Station Amazon",
    "Distance (km)",
    "Tel (à remplir)",
    "Email pro (à remplir)",
    "Détail score",
    "URL Annuaire",
    "Recherche Tel (Google)",
    "LinkedIn société",
    "LinkedIn dirigeant",
    "Google Maps siège",
    "MAJ INSEE",
    "Notes Ousmane",
]

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
COL_WIDTHS = {
    "A": 6, "B": 13, "C": 12, "D": 36, "E": 8, "F": 12, "G": 6, "H": 10,
    "I": 6, "J": 6, "K": 14, "L": 16, "M": 24, "N": 9, "O": 7, "P": 14,
    "Q": 30, "R": 45, "S": 8, "T": 18, "U": 6, "V": 11, "W": 11, "X": 9,
    "Y": 9, "Z": 18, "AA": 28, "AB": 70, "AC": 55,
    "AD": 28, "AE": 28, "AE": 28, "AF": 28, "AG": 22, "AH": 12, "AI": 30,
}


def style_header(ws) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "C2"
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w


def _quote_url(s: str) -> str:
    """URL-quote pour Google search."""
    from urllib.parse import quote_plus
    return quote_plus(s)


def company_row(c: Company) -> list[Any]:
    approche = ""
    if c.dirigeant_age is not None:
        approche = "Tutoyer (V1)" if c.dirigeant_age < 45 else "Vouvoyer (V2)"

    annuaire_url = f"https://annuaire-entreprises.data.gouv.fr/entreprise/{c.siren}"

    nom_query = f'"{c.nom}" "{c.commune}" telephone' if c.commune else f'"{c.nom}" telephone'
    google_tel_url = f"https://www.google.com/search?q={_quote_url(nom_query)}"

    li_company_q = f'site:linkedin.com/company "{c.nom}"'
    linkedin_company_url = f"https://www.google.com/search?q={_quote_url(li_company_q)}"

    if c.dirigeant_prenom and c.dirigeant_nom:
        li_dirigeant_q = f'site:linkedin.com/in "{c.dirigeant_prenom} {c.dirigeant_nom}" "{c.commune}"'
        linkedin_dirigeant_url = f"https://www.google.com/search?q={_quote_url(li_dirigeant_q)}"
    else:
        linkedin_dirigeant_url = ""

    if c.latitude and c.longitude:
        gmaps_url = f"https://www.google.com/maps/search/?api=1&query={c.latitude},{c.longitude}"
    else:
        gmaps_url = f"https://www.google.com/maps/search/?api=1&query={_quote_url(c.adresse)}"

    return [
        c.score,
        "Lead",
        c.siren,
        c.nom,
        c.forme_juridique_libelle,
        c.date_creation,
        c.age_annees,
        c.tranche_effectif_libelle,
        c.categorie_entreprise,
        c.nb_etablissements,
        c.dirigeant_prenom,
        c.dirigeant_nom,
        c.dirigeant_fonction,
        c.dirigeant_annee_naissance,
        c.dirigeant_age,
        approche,
        " | ".join(c.other_dirigeants_pp),
        c.adresse,
        c.code_postal,
        c.commune,
        c.departement,
        c.latitude,
        c.longitude,
        c.nearest_station_code,
        c.nearest_station_km if c.nearest_station_km < 999 else None,
        "",  # tel à remplir manuellement
        "",  # email à remplir
        " ; ".join(c.score_details),
        annuaire_url,
        google_tel_url,
        linkedin_company_url,
        linkedin_dirigeant_url,
        gmaps_url,
        c.date_maj_insee,
        "",  # notes
    ]


URL_COL_INDICES = [29, 30, 31, 32, 33]  # 1-indexed: URL Annuaire, Google Tel, LinkedIn société, LinkedIn dirigeant, Maps
LINK_FONT = Font(color="2563EB", underline="single", size=11)


def fill_sheet(ws, companies: list[Company], color_mode: str = "gradient") -> None:
    ws.append(HEADERS)
    for c in companies:
        ws.append(company_row(c))
    style_header(ws)

    # Convert URL columns to clickable hyperlinks with friendly label
    label_map = {
        29: "Annuaire",
        30: "Search tel",
        31: "LinkedIn société",
        32: "LinkedIn dirigeant",
        33: "Maps",
    }
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in URL_COL_INDICES:
            cell = ws.cell(row=row_idx, column=col_idx)
            url = cell.value
            if url and isinstance(url, str) and url.startswith("http"):
                cell.hyperlink = url
                cell.value = label_map.get(col_idx, "Lien")
                cell.font = LINK_FONT

    # Row coloring par score
    fill_t1 = PatternFill(start_color="A7F3D0", end_color="A7F3D0", fill_type="solid")  # vert vif
    fill_t2 = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")  # vert clair
    fill_t3 = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # jaune
    fill_t4 = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")  # rouge clair
    if color_mode == "gradient":
        for row_idx in range(2, ws.max_row + 1):
            sv = ws.cell(row=row_idx, column=1).value or 0
            try:
                sv = int(sv)
            except (TypeError, ValueError):
                sv = 0
            fill = None
            if sv >= 14:
                fill = fill_t1
            elif sv >= 11:
                fill = fill_t2
            elif sv >= 8:
                fill = fill_t3
            elif sv >= 5:
                fill = fill_t4
            if fill:
                for col_idx in range(1, len(HEADERS) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = fill
                    # Re-apply link font on top of fill for URL cols
                    if col_idx in URL_COL_INDICES and cell.hyperlink:
                        cell.font = LINK_FONT


def write_excel(companies: list[Company], out_path: Path) -> None:
    wb = Workbook()

    tier1 = [c for c in companies if c.score >= 14]
    tier2 = [c for c in companies if 11 <= c.score < 14]
    tier3 = [c for c in companies if 8 <= c.score < 11]
    tier4 = [c for c in companies if 5 <= c.score < 8]
    tier_other = [c for c in companies if c.score < 5]

    # === Sheet 1 : Action immédiate (Tier 1, score >= 14) ===
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = "Tier 1 - Action immédiate"
    fill_sheet(ws, tier1)

    # === Sheet 2 : Tier 2 (11-13) — high priority ===
    ws2 = wb.create_sheet("Tier 2 - High priority")
    fill_sheet(ws2, tier2)

    # === Sheet 3 : Tier 3 (8-10) — secondary ===
    ws3 = wb.create_sheet("Tier 3 - Secondary")
    fill_sheet(ws3, tier3)

    # === Sheet 4 : Tier 4 (5-7) — exploration ===
    ws4 = wb.create_sheet("Tier 4 - Exploration")
    fill_sheet(ws4, tier4)

    # === Sheet 5 : All — vue complète triée ===
    ws_all = wb.create_sheet("All - Vue complète")
    fill_sheet(ws_all, companies)

    # === Sheet 6 : Stations Amazon ===
    ws_st = wb.create_sheet("Stations Amazon")
    ws_st.append(["Code station", "Localisation", "Latitude", "Longitude"])
    for code, (label, lat, lng) in AMAZON_STATIONS.items():
        ws_st.append([code, label, lat, lng])
    for cell in ws_st[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws_st.column_dimensions["A"].width = 12
    ws_st.column_dimensions["B"].width = 45
    ws_st.column_dimensions["C"].width = 12
    ws_st.column_dimensions["D"].width = 12

    # === Sheet 7 : Méthodologie ===
    ws_meth = wb.create_sheet("Méthodologie")
    info = [
        ["DSPilot — Leads DSP Amazon France Île-de-France"],
        [""],
        ["Source données", "Annuaire Entreprises (recherche-entreprises.api.gouv.fr) — État FR"],
        ["Date extraction", time.strftime("%Y-%m-%d %H:%M")],
        ["Auteur", "Ousmane Sané — DSPilot — pilote distribution IDF"],
        ["Total leads sourcés", f"{len(companies)} sociétés (SIREN uniques)"],
        [""],
        ["Filtres appliqués"],
        ["", "APE / NAF principal", "49.41B (Transports routiers de fret de proximité)"],
        ["", "Départements", "75, 77, 78, 91, 92, 93, 94, 95 (Île-de-France)"],
        ["", "État administratif", "Actif"],
        ["", "Effectif salarié", "10 à 199 (tranches INSEE 11, 12, 21, 22)"],
        ["", "Catégorie", "PME / ETI uniquement (GE = grandes entreprises exclues)"],
        ["", "Dirigeants retenus", "Personne physique (commissaires aux comptes exclus)"],
        [""],
        ["Système de scoring (max 20 pts)"],
        ["", "Proximité station Amazon"],
        ["", "  +6", "Siège <= 2 km d'une station"],
        ["", "  +4", "Siège <= 5 km"],
        ["", "  +2", "Siège <= 10 km"],
        ["", "  +1", "Siège <= 20 km"],
        ["", "Cohérence programme DSP Amazon FR (lancé 2018)"],
        ["", "  +3", "Création >= 2018"],
        ["", "  +1", "Création 2015-2017"],
        ["", "Effectif (DSP type = 30-150 livreurs)"],
        ["", "  +4", "50-99 salariés (sweet spot)"],
        ["", "  +3", "20-49 salariés"],
        ["", "  +1", "10-19 ou 100-199"],
        ["", "Forme juridique"],
        ["", "  +2", "SAS / SARL / SASU (le programme DSP exige une société)"],
        ["", "Dirigeant"],
        ["", "  +2", "Dirigeant nominatif identifié (contact direct possible)"],
        ["", "Structure"],
        ["", "  +2", "Mono-établissement (vrai DSP indépendant)"],
        ["", "  +1", "2-3 établissements"],
        [""],
        ["Lecture des tiers"],
        ["", "Tier 1 (>= 14)", f"🟢 {len(tier1)} sociétés — Action immédiate (top priorité, appeler en 1er)"],
        ["", "Tier 2 (11-13)", f"🟢 {len(tier2)} sociétés — High priority"],
        ["", "Tier 3 (8-10)", f"🟡 {len(tier3)} sociétés — Secondary, vérifier manuellement"],
        ["", "Tier 4 (5-7)", f"🟠 {len(tier4)} sociétés — Exploration, probablement non-DSP"],
        ["", "Other (< 5)", f"⚪ {len(tier_other)} sociétés — Bruit (transporteurs classiques)"],
        [""],
        ["Workflow recommandé"],
        ["", "1.", "Parcourir Tier 1 — pour chaque ligne, vérifier réseau Ousmane (col 'Notes Ousmane')"],
        ["", "2.", "Enrichir tel (Pages Jaunes) + email (Hunter.io free 25/mois) → cols Tel / Email"],
        ["", "3.", "Appel direct 5/jour selon Approche (V1 tutoyer si <45 ans, V2 vouvoyer si 50+)"],
        ["", "4.", "Update col Stage : Lead → Enriched → Contacté → Booked → Démo → Closed won/lost"],
        ["", "5.", "Si Tier 1 épuisé sans conversion, passer Tier 2"],
        [""],
        ["Note RGPD", "Toutes les données sont publiques (INSEE SIRENE + RNE — État FR). Aucune adresse personnelle de dirigeant n'est stockée — uniquement l'adresse du siège social. Conforme RGPD pour outreach B2B."],
    ]
    for row in info:
        ws_meth.append(row)
    ws_meth.column_dimensions["A"].width = 6
    ws_meth.column_dimensions["B"].width = 24
    ws_meth.column_dimensions["C"].width = 90
    ws_meth["A1"].font = Font(bold=True, size=14, color="1F2937")
    for r_idx, row in enumerate(info, start=1):
        if row and row[0] and row[0] not in ("", " ") and not row[0].startswith(" "):
            if len(row) >= 2 and row[1] and not str(row[1]).startswith(" "):
                continue
        if not row:
            continue
        first = row[0] if row else ""
        if first in ("Filtres appliqués", "Système de scoring (max 20 pts)", "Lecture des tiers", "Workflow recommandé"):
            ws_meth.cell(row=r_idx, column=1).font = Font(bold=True, size=12, color="1F2937")

    wb.save(out_path)


def write_html_map(companies: list[Company], out_path: Path) -> None:
    """Map interactive Leaflet.js + filtres + popups détaillés."""
    leads_data = []
    for c in companies:
        if c.latitude is None or c.longitude is None:
            continue
        nom_query = f'"{c.nom}" "{c.commune}" telephone' if c.commune else f'"{c.nom}" telephone'
        gtel = f"https://www.google.com/search?q={_quote_url(nom_query)}"
        li_co = f"https://www.google.com/search?q={_quote_url('site:linkedin.com/company ' + chr(34) + c.nom + chr(34))}"
        li_dir = ""
        if c.dirigeant_prenom and c.dirigeant_nom:
            li_dir = f"https://www.google.com/search?q={_quote_url(f'site:linkedin.com/in {chr(34)}{c.dirigeant_prenom} {c.dirigeant_nom}{chr(34)} {chr(34)}{c.commune}{chr(34)}')}"
        gmaps = f"https://www.google.com/maps/search/?api=1&query={c.latitude},{c.longitude}"
        annuaire = f"https://annuaire-entreprises.data.gouv.fr/entreprise/{c.siren}"

        approche = ""
        if c.dirigeant_age is not None:
            approche = "Tutoyer V1" if c.dirigeant_age < 45 else "Vouvoyer V2"

        # Tier
        if c.score >= 14:
            tier = 1
        elif c.score >= 11:
            tier = 2
        elif c.score >= 8:
            tier = 3
        elif c.score >= 5:
            tier = 4
        else:
            tier = 5

        leads_data.append({
            "siren": c.siren,
            "nom": c.nom,
            "lat": c.latitude,
            "lng": c.longitude,
            "score": c.score,
            "tier": tier,
            "forme": c.forme_juridique_libelle,
            "creation": c.date_creation,
            "age": c.age_annees,
            "effectif": c.tranche_effectif_libelle,
            "etabts": c.nb_etablissements,
            "dirigeant": f"{c.dirigeant_prenom} {c.dirigeant_nom}".strip(),
            "fonction": c.dirigeant_fonction,
            "age_dir": c.dirigeant_age,
            "approche": approche,
            "adresse": c.adresse,
            "cp": c.code_postal,
            "commune": c.commune,
            "dept": c.departement,
            "station": c.nearest_station_code,
            "station_label": c.nearest_station_label,
            "km": round(c.nearest_station_km, 2) if c.nearest_station_km < 999 else None,
            "score_details": " ; ".join(c.score_details),
            "gtel": gtel,
            "li_co": li_co,
            "li_dir": li_dir,
            "gmaps": gmaps,
            "annuaire": annuaire,
        })

    stations_data = [
        {"code": code, "label": label, "lat": lat, "lng": lng}
        for code, (label, lat, lng) in AMAZON_STATIONS.items()
    ]

    leads_json = json.dumps(leads_data, ensure_ascii=False)
    stations_json = json.dumps(stations_data, ensure_ascii=False)

    html = """<!doctype html>
<html lang="fr">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>DSPilot — Carte leads DSP IDF</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css">
<link rel="stylesheet" href="https://unpkg.com/leaflet.markercluster@1.5.3/dist/MarkerCluster.css">
<link rel="stylesheet" href="https://unpkg.com/leaflet.markercluster@1.5.3/dist/MarkerCluster.Default.css">
<style>
  * { box-sizing: border-box; }
  body { margin:0; font:14px/1.4 -apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif; color:#111827; }
  #app { display:grid; grid-template-columns:340px 1fr; height:100vh; }
  #sidebar { background:#0f172a; color:#e2e8f0; padding:16px; overflow-y:auto; }
  #sidebar h1 { margin:0 0 4px; font-size:18px; color:#f8fafc; }
  #sidebar .meta { color:#94a3b8; font-size:12px; margin-bottom:16px; }
  #sidebar h2 { margin:18px 0 8px; font-size:13px; text-transform:uppercase; letter-spacing:.05em; color:#94a3b8; font-weight:600; }
  #sidebar .stat { display:flex; justify-content:space-between; padding:6px 0; border-bottom:1px solid #1e293b; font-size:13px; }
  #sidebar .stat .v { font-weight:600; color:#f8fafc; }
  #sidebar input[type=text], #sidebar select { width:100%; padding:7px 10px; border-radius:6px; border:1px solid #334155; background:#1e293b; color:#f1f5f9; font-size:13px; margin-bottom:8px; }
  #sidebar label { display:flex; align-items:center; gap:6px; padding:4px 0; font-size:13px; cursor:pointer; }
  #sidebar label input { margin:0; }
  #sidebar .tier-pill { display:inline-block; width:10px; height:10px; border-radius:50%; margin-right:6px; vertical-align:middle; }
  #sidebar .reset { display:block; width:100%; padding:8px; margin-top:12px; background:#1e293b; color:#e2e8f0; border:1px solid #334155; border-radius:6px; cursor:pointer; font-size:13px; }
  #sidebar .reset:hover { background:#334155; }
  #map { height:100vh; }
  .popup { font-size:13px; min-width:280px; }
  .popup h3 { margin:0 0 4px; font-size:15px; color:#0f172a; }
  .popup .siren { color:#64748b; font-size:11px; font-family:ui-monospace,monospace; }
  .popup .score { display:inline-block; padding:2px 8px; border-radius:4px; font-weight:700; margin-left:8px; }
  .popup .score.t1 { background:#10b981; color:white; }
  .popup .score.t2 { background:#34d399; color:white; }
  .popup .score.t3 { background:#fbbf24; color:#78350f; }
  .popup .score.t4 { background:#f87171; color:white; }
  .popup .score.t5 { background:#cbd5e1; color:#1e293b; }
  .popup .grid { display:grid; grid-template-columns:auto 1fr; gap:4px 12px; margin:8px 0; }
  .popup .grid b { color:#475569; font-weight:500; }
  .popup .links { display:flex; flex-wrap:wrap; gap:6px; margin-top:10px; padding-top:10px; border-top:1px solid #e2e8f0; }
  .popup .links a { display:inline-block; padding:5px 10px; background:#2563eb; color:white; text-decoration:none; border-radius:5px; font-size:12px; font-weight:500; }
  .popup .links a:hover { background:#1d4ed8; }
  .popup .links a.secondary { background:#64748b; }
  .popup .links a.secondary:hover { background:#475569; }
  .leaflet-popup-content { margin:14px; }
  .station-icon { background:#a855f7; border:3px solid white; border-radius:50%; width:18px; height:18px; box-shadow:0 0 0 2px #a855f7; }
  .station-icon::after { content:""; }
</style>
</head>
<body>
<div id="app">
  <div id="sidebar">
    <h1>DSPilot — Leads DSP IDF</h1>
    <div class="meta">__GENERATED__ • __TOTAL__ leads • Source : Annuaire Entreprises</div>

    <h2>Recherche</h2>
    <input type="text" id="search" placeholder="Société, dirigeant, ville, SIREN…">

    <h2>Tiers (score DSP-fit)</h2>
    <label><input type="checkbox" class="tier" value="1" checked><span class="tier-pill" style="background:#10b981"></span>Tier 1 — Action immédiate (≥14)</label>
    <label><input type="checkbox" class="tier" value="2" checked><span class="tier-pill" style="background:#34d399"></span>Tier 2 — High priority (11-13)</label>
    <label><input type="checkbox" class="tier" value="3"><span class="tier-pill" style="background:#fbbf24"></span>Tier 3 — Secondary (8-10)</label>
    <label><input type="checkbox" class="tier" value="4"><span class="tier-pill" style="background:#f87171"></span>Tier 4 — Exploration (5-7)</label>
    <label><input type="checkbox" class="tier" value="5"><span class="tier-pill" style="background:#cbd5e1"></span>Bruit (&lt;5)</label>

    <h2>Station Amazon</h2>
    <select id="station-filter"><option value="">Toutes les stations</option></select>

    <h2>Effectif</h2>
    <select id="effectif-filter">
      <option value="">Tous</option>
      <option value="10-19">10-19</option>
      <option value="20-49">20-49</option>
      <option value="50-99">50-99</option>
      <option value="100-199">100-199</option>
    </select>

    <h2>Département</h2>
    <select id="dept-filter">
      <option value="">Tous</option>
      <option value="75">75 Paris</option>
      <option value="77">77 Seine-et-Marne</option>
      <option value="78">78 Yvelines</option>
      <option value="91">91 Essonne</option>
      <option value="92">92 Hauts-de-Seine</option>
      <option value="93">93 Seine-Saint-Denis</option>
      <option value="94">94 Val-de-Marne</option>
      <option value="95">95 Val-d'Oise</option>
    </select>

    <h2 style="margin-top:18px">Visibles</h2>
    <div class="stat"><span>Leads affichés</span><span class="v" id="count-visible">0</span></div>
    <div class="stat"><span>Tier 1 visibles</span><span class="v" id="count-t1">0</span></div>
    <div class="stat"><span>Tier 2 visibles</span><span class="v" id="count-t2">0</span></div>

    <button class="reset" onclick="resetFilters()">Reset filtres</button>
  </div>

  <div id="map"></div>
</div>

<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<script src="https://unpkg.com/leaflet.markercluster@1.5.3/dist/leaflet.markercluster.js"></script>
<script>
const LEADS = __LEADS_JSON__;
const STATIONS = __STATIONS_JSON__;

const tierColor = {1:"#10b981",2:"#34d399",3:"#fbbf24",4:"#f87171",5:"#cbd5e1"};
const tierBorder = {1:"#065f46",2:"#047857",3:"#b45309",4:"#991b1b",5:"#475569"};

const map = L.map("map", { preferCanvas:true }).setView([48.8566, 2.3522], 10);
L.tileLayer("https://{s}.basemaps.cartocdn.com/light_all/{z}/{x}/{y}{r}.png", {
  attribution:"&copy; OpenStreetMap & CARTO",
  maxZoom:19
}).addTo(map);

// Stations Amazon = pins violets fixes
STATIONS.forEach(s => {
  L.marker([s.lat, s.lng], {
    icon: L.divIcon({
      className:"",
      html:`<div style="background:#a855f7;border:3px solid white;border-radius:50%;width:22px;height:22px;box-shadow:0 0 0 2px #a855f7,0 1px 4px rgba(0,0,0,.3)"></div>`,
      iconSize:[22,22], iconAnchor:[11,11]
    }),
    zIndexOffset: 1000
  }).addTo(map).bindPopup(`<b>Station Amazon ${s.code}</b><br>${s.label}`);
});

// Build station-filter select
const stationFilter = document.getElementById("station-filter");
[...new Set(LEADS.map(l => l.station).filter(Boolean))].sort().forEach(code => {
  const opt = document.createElement("option");
  opt.value = code;
  opt.textContent = code;
  stationFilter.appendChild(opt);
});

const cluster = L.markerClusterGroup({
  maxClusterRadius: 50,
  spiderfyOnMaxZoom: true,
  showCoverageOnHover: false,
});
const markers = [];

LEADS.forEach(l => {
  const color = tierColor[l.tier] || "#cbd5e1";
  const border = tierBorder[l.tier] || "#475569";
  const size = l.tier === 1 ? 14 : l.tier === 2 ? 12 : 10;
  const m = L.circleMarker([l.lat, l.lng], {
    radius: size/2,
    fillColor: color,
    color: border,
    weight: 2,
    fillOpacity: 0.85,
  });
  m.leadData = l;
  m.bindPopup(() => buildPopup(l), { maxWidth: 360 });
  markers.push(m);
  cluster.addLayer(m);
});
map.addLayer(cluster);

function buildPopup(l) {
  const dirig = l.dirigeant
    ? `<div class="grid"><b>Dirigeant</b><span>${escapeHtml(l.dirigeant)}</span><b>Fonction</b><span>${escapeHtml(l.fonction || "")}</span><b>Âge</b><span>${l.age_dir ?? "?"} ans → <b>${l.approche || ""}</b></span></div>`
    : `<div class="grid"><b>Dirigeant</b><span style="color:#94a3b8">Non identifié</span></div>`;
  const km = l.km != null ? `${l.km} km` : "?";
  const links = [
    `<a href="${l.gtel}" target="_blank">📞 Trouver tel</a>`,
    `<a href="${l.li_co}" target="_blank">LinkedIn société</a>`,
    l.li_dir ? `<a href="${l.li_dir}" target="_blank">LinkedIn dirigeant</a>` : "",
    `<a href="${l.gmaps}" target="_blank" class="secondary">Google Maps</a>`,
    `<a href="${l.annuaire}" target="_blank" class="secondary">Annuaire</a>`,
  ].filter(Boolean).join("");
  return `<div class="popup">
    <h3>${escapeHtml(l.nom)} <span class="score t${l.tier}">${l.score}</span></h3>
    <div class="siren">SIREN ${l.siren} • ${l.forme || "?"} • Créée ${l.creation || "?"} (${l.age || "?"} ans) • ${l.etabts || 1} étabt(s)</div>
    <div class="grid" style="margin-top:8px"><b>Effectif</b><span>${l.effectif || "?"} salariés</span><b>Adresse</b><span>${escapeHtml(l.adresse)}</span><b>Station</b><span>${l.station || "?"} (${km})</span></div>
    ${dirig}
    <div class="grid"><b>Score detail</b><span style="font-size:11px;color:#64748b">${escapeHtml(l.score_details || "")}</span></div>
    <div class="links">${links}</div>
  </div>`;
}

function escapeHtml(s) { return String(s ?? "").replace(/[&<>"']/g, c => ({"&":"&amp;","<":"&lt;",">":"&gt;","\\"":"&quot;","'":"&#39;"})[c]); }

const filters = { tiers: new Set([1,2]), search: "", station: "", effectif: "", dept: "" };

function applyFilters() {
  const q = filters.search.toLowerCase();
  cluster.clearLayers();
  let visible = 0, t1 = 0, t2 = 0;
  markers.forEach(m => {
    const l = m.leadData;
    if (!filters.tiers.has(l.tier)) return;
    if (filters.station && l.station !== filters.station) return;
    if (filters.effectif && l.effectif !== filters.effectif) return;
    if (filters.dept && l.dept !== filters.dept) return;
    if (q) {
      const hay = `${l.nom} ${l.dirigeant} ${l.commune} ${l.siren} ${l.adresse}`.toLowerCase();
      if (!hay.includes(q)) return;
    }
    cluster.addLayer(m);
    visible++;
    if (l.tier === 1) t1++;
    if (l.tier === 2) t2++;
  });
  document.getElementById("count-visible").textContent = visible;
  document.getElementById("count-t1").textContent = t1;
  document.getElementById("count-t2").textContent = t2;
}

document.querySelectorAll(".tier").forEach(cb => cb.addEventListener("change", e => {
  const v = parseInt(cb.value, 10);
  if (cb.checked) filters.tiers.add(v); else filters.tiers.delete(v);
  applyFilters();
}));
document.getElementById("search").addEventListener("input", e => { filters.search = e.target.value; applyFilters(); });
document.getElementById("station-filter").addEventListener("change", e => { filters.station = e.target.value; applyFilters(); });
document.getElementById("effectif-filter").addEventListener("change", e => { filters.effectif = e.target.value; applyFilters(); });
document.getElementById("dept-filter").addEventListener("change", e => { filters.dept = e.target.value; applyFilters(); });

function resetFilters() {
  filters.tiers = new Set([1,2]);
  filters.search = ""; filters.station = ""; filters.effectif = ""; filters.dept = "";
  document.querySelectorAll(".tier").forEach(cb => { cb.checked = [1,2].includes(parseInt(cb.value,10)); });
  document.getElementById("search").value = "";
  document.getElementById("station-filter").value = "";
  document.getElementById("effectif-filter").value = "";
  document.getElementById("dept-filter").value = "";
  applyFilters();
}

applyFilters();
</script>
</body>
</html>"""

    html = (
        html
        .replace("__GENERATED__", time.strftime("%Y-%m-%d %H:%M"))
        .replace("__TOTAL__", str(len(leads_data)))
        .replace("__LEADS_JSON__", leads_json)
        .replace("__STATIONS_JSON__", stations_json)
    )

    out_path.write_text(html, encoding="utf-8")


def main() -> int:
    print(f"[1/5] Pull APE 49.41B effectif 10-199 actif sur 8 départements IDF…", file=sys.stderr)
    raw_companies: list[dict[str, Any]] = []
    for dept in DEPARTEMENTS_IDF:
        n_before = len(raw_companies)
        for raw in fetch_dept(dept):
            raw_companies.append(raw)
        print(f"  ✓ Dept {dept} : {len(raw_companies) - n_before} sociétés brutes", file=sys.stderr)

    print(f"[2/5] {len(raw_companies)} brutes — dédoublonnage par SIREN…", file=sys.stderr)
    seen_sirens: set[str] = set()
    unique_raws: list[dict[str, Any]] = []
    for raw in raw_companies:
        siren = raw.get("siren")
        if siren and siren not in seen_sirens:
            seen_sirens.add(siren)
            unique_raws.append(raw)
    print(f"  ✓ {len(unique_raws)} SIREN uniques", file=sys.stderr)

    print(f"[3/5] Parsing + filtrage GE (grandes entreprises)…", file=sys.stderr)
    companies: list[Company] = []
    for raw in unique_raws:
        c = parse_company(raw)
        if c is not None:
            companies.append(c)
    print(f"  ✓ {len(companies)} après filtrage", file=sys.stderr)

    print(f"[4/5] Scoring DSP candidate…", file=sys.stderr)
    companies = [score_company(c) for c in companies]
    companies.sort(key=lambda c: (-c.score, c.nearest_station_km))

    tier_top = sum(1 for c in companies if c.score >= 14)
    tier_high = sum(1 for c in companies if 11 <= c.score < 14)
    tier_med = sum(1 for c in companies if 8 <= c.score < 11)
    tier_low = sum(1 for c in companies if 5 <= c.score < 8)
    print(
        f"  ✓ Tier 1 (>=14) : {tier_top} "
        f"| Tier 2 (11-13) : {tier_high} "
        f"| Tier 3 (8-10) : {tier_med} "
        f"| Tier 4 (5-7) : {tier_low}",
        file=sys.stderr,
    )

    out_dir = Path("/Users/ousmane/wiki/wiki")
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "leads-DSPilot-IDF.xlsx"
    print(f"[5/5] Écriture Excel → {out_path}", file=sys.stderr)
    write_excel(companies, out_path)
    print(f"  ✓ Done. {out_path}", file=sys.stderr)

    # Top 10 récap stdout
    print("\n=== TOP 10 leads (score desc) ===")
    for c in companies[:10]:
        dirigeant = (
            f"{c.dirigeant_prenom} {c.dirigeant_nom} ({c.dirigeant_fonction})"
            if c.dirigeant_prenom
            else "(dirigeant non identifié)"
        )
        print(
            f"[{c.score:2d}] {c.nom[:35]:35s} | {c.commune[:18]:18s} | "
            f"{c.tranche_effectif_libelle:8s} | {c.nearest_station_code:6s} {c.nearest_station_km:5.1f}km | {dirigeant}"
        )
    return 0


if __name__ == "__main__":
    sys.exit(main())
